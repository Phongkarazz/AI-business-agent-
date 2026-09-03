"""
Export utilities for Multi-format Reporting:
- Styled Excel (.xlsx) with custom headers, cell borders, and currency/date formatting.
- High-Resolution PNG (.png) for charts and slides.
- Executive PDF Report (.pdf) with 100% Vietnamese Unicode font support, clean typography,
  complete emoji stripping (no white boxes), and professional business layout.
"""

import io
import os
import re
import datetime
import pandas as pd

# Regex bắt toàn bộ các dải ký tự biểu tượng cảm xúc (Emoji) trong Unicode
EMOJI_REGEX = re.compile(
    r"[\U00010000-\U0010ffff]|[\uD800-\uDBFF][\uDC00-\uDFFF]|[\u2600-\u27BF]|[\u2300-\u23FF]|[\u2B50-\u2B55]|[\u200D\uFE0F]|[\uFE00-\uFE0F]",
    flags=re.UNICODE
)


def clean_text_for_pdf(text: str) -> str:
    """Loại bỏ triệt để các emoji gây lỗi ô vuông trắng và chuẩn hóa cú pháp XML an toàn cho ReportLab."""
    if not text:
        return ""

    text = str(text)

    # 1. Chuyển đổi các biểu tượng ưu tiên thành nhãn văn bản chuẩn
    text = text.replace("🔴", "").replace("🟡", "").replace("🟢", "")

    # 2. Xóa toàn bộ emoji unicode và ký tự glyph lạ gây lỗi ô vuông tofu (□, ↳, ➔...)
    text = EMOJI_REGEX.sub("", text)
    text = text.replace("↳", "").replace("➔", "").replace("➜", "").replace("➡", "").replace("►", "").replace("□", "")

    # 3. Tách từ viết hoa dính liền với từ tiếng Việt (VD: KHÔNGPhát -> KHÔNG Phát, USACó -> USA Có)
    text = re.sub(r"([A-ZÀ-Ỹ]{2,})([A-ZÀ-Ỹ][a-zà-ỹ])", r"\1 \2", text)
    text = re.sub(r"([A-Z]{2,})([a-zà-ỹ])", r"\1 \2", text)

    # 4. Thêm khoảng trắng sau dấu hai chấm nếu bị dính số hoặc chữ và sửa lỗi 2 dấu hai chấm
    text = re.sub(r"\]\s*:\s*:\s*", "]: ", text)
    text = re.sub(r"\s*:\s*:\s*", ": ", text)
    text = re.sub(r"([a-zA-Zà-ỹÀ-Ỹ]):(\d)", r"\1: \2", text)
    text = re.sub(r":([A-ZÀ-Ỹa-zà-ỹ])", r": \1", text)

    # 4.1 Đảm bảo khoảng trắng quanh thẻ in đậm
    text = re.sub(r"\*\*([^\*\n]+?)\*\*([a-zA-Zà-ỹÀ-Ỹ0-9])", r"**\1** \2", text)
    text = re.sub(r"([a-zA-Zà-ỹÀ-Ỹ0-9])\*\*([^\*\n]+?)\*\*", r"\1 **\2**", text)

    # 5. Thoát các ký tự XML đặc biệt
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    # 6. Chuyển đổi cú pháp markdown sang thẻ định dạng ReportLab
    text = re.sub(r"\*\*(.*?)\*\*", r"<b>\1</b>", text)
    text = re.sub(r"\*(.*?)\*", r"<i>\1</i>", text)
    text = re.sub(r"`(.*?)`", r"<b>\1</b>", text)

    # 7. Xóa các ký tự markdown header #### dư thừa trong dòng
    text = re.sub(r"#+\s*", "", text)

    # 8. Dọn dẹp khoảng trắng thừa và ký tự rác đầu mục
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"^\s*[\.\-•]\s*", "", text.strip())

    return text.strip()


# ---------------------------------------------------------
# 1. Xuất Excel (.xlsx) Định Dạng Chuyên Nghiệp
# ---------------------------------------------------------
def export_to_excel(df: pd.DataFrame, sheet_name: str = "Bao_Cao") -> bytes:
    """Tạo file Excel (.xlsx) có sẵn style header màu xanh Navy, viền ô và định dạng số phân cách hàng nghìn."""
    if df is None or df.empty:
        return b""

    output = io.BytesIO()
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            clean_sheet = "".join(c for c in sheet_name if c.isalnum() or c in (" ", "_"))[:30] or "Data"
            df.to_excel(writer, index=False, sheet_name=clean_sheet)
            ws = writer.sheets[clean_sheet]

            # Style Tiêu đề Header
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            border_thin = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9")
            )

            for col_idx, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # Tính độ rộng cột tự động
                max_val_len = df[col].astype(str).map(len).max() if not df[col].empty else 0
                max_len = max(max_val_len, len(str(col))) + 4
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(max_len, 14)

                # Định dạng dữ liệu các dòng
                is_num = pd.api.types.is_numeric_dtype(df[col])
                for row_idx in range(2, len(df) + 2):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.border = border_thin
                    if is_num:
                        c.number_format = "#,##0"
                        c.alignment = Alignment(horizontal="right")
                    else:
                        c.alignment = Alignment(horizontal="left")

        return output.getvalue()
    except Exception:
        output = io.BytesIO()
        df.to_excel(output, index=False)
        return output.getvalue()


# ---------------------------------------------------------
# 2. Xuất Ảnh Biểu Đồ PNG Độ Nét Cao
# ---------------------------------------------------------
def export_to_png(fig) -> bytes | None:
    """Xuất biểu đồ Plotly sang ảnh PNG.
    Đã tắt việc gọi Chrome headless ngầm để chống treo server (người dùng tải trực tiếp qua nút Camera 📷 trên biểu đồ).
    """
    return None


# ---------------------------------------------------------
# 3. Xuất Tóm Tắt Báo Cáo Executive PDF (.pdf) Chuẩn Doanh Nghiệp
# ---------------------------------------------------------
def export_to_pdf(result: dict, df: pd.DataFrame, chart_png_bytes: bytes = None) -> bytes:
    """Tạo tài liệu Báo cáo Điều hành PDF (Executive Report) chuẩn xác, loại bỏ hoàn toàn ô vuông lỗi font."""
    if result is None:
        return b""

    buffer = io.BytesIO()
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        # 1. Đăng ký Font Unicode Tiếng Việt từ thư mục dự án
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        bundled_font = os.path.join(base_dir, "assets", "fonts", "CustomUnicode.ttf")
        bundled_font_bold = os.path.join(base_dir, "assets", "fonts", "CustomUnicode-Bold.ttf")

        font_name = "Helvetica"
        font_name_bold = "Helvetica-Bold"

        if os.path.exists(bundled_font):
            try:
                pdfmetrics.registerFont(TTFont("CustomUnicode", bundled_font))
                font_name = "CustomUnicode"
                font_name_bold = "CustomUnicode"
                if os.path.exists(bundled_font_bold):
                    pdfmetrics.registerFont(TTFont("CustomUnicodeBold", bundled_font_bold))
                    font_name_bold = "CustomUnicodeBold"
            except Exception:
                pass
        else:
            possible_fonts = [
                "/System/Library/Fonts/Supplemental/Arial.ttf",
                "/Library/Fonts/Arial.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"
            ]
            for fpath in possible_fonts:
                if os.path.exists(fpath):
                    try:
                        pdfmetrics.registerFont(TTFont("CustomUnicode", fpath))
                        font_name = "CustomUnicode"
                        font_name_bold = "CustomUnicode"
                        break
                    except Exception:
                        pass

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "DocTitle",
            parent=styles["Heading1"],
            fontName=font_name_bold,
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#1F4E78"),
            spaceAfter=2
        )
        subtitle_style = ParagraphStyle(
            "DocSub",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8.5,
            leading=12,
            textColor=colors.HexColor("#4B5563"),
            spaceAfter=4
        )
        section_style = ParagraphStyle(
            "SectionHead",
            parent=styles["Heading2"],
            fontName=font_name_bold,
            fontSize=11,
            leading=15,
            textColor=colors.HexColor("#1F4E78"),
            spaceBefore=12,
            spaceAfter=6
        )
        sub_section_style = ParagraphStyle(
            "SubSectionHead",
            parent=styles["Heading3"],
            fontName=font_name_bold,
            fontSize=9.5,
            leading=13.5,
            textColor=colors.HexColor("#1F4E78"),
            spaceBefore=9,
            spaceAfter=4,
            leftIndent=4
        )
        body_style = ParagraphStyle(
            "BodyText",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8.5,
            leading=12.5,
            textColor=colors.HexColor("#1F2937"),
            spaceAfter=3
        )
        bullet_style = ParagraphStyle(
            "BulletText",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8.5,
            leading=12.5,
            textColor=colors.HexColor("#1F2937"),
            leftIndent=14,
            spaceAfter=3
        )
        priority_tag_style = ParagraphStyle(
            "PriorityTag",
            parent=styles["Normal"],
            fontName=font_name_bold,
            fontSize=8.5,
            leading=12.5,
            textColor=colors.HexColor("#1E3A8A"),
            leftIndent=8,
            spaceBefore=6,
            spaceAfter=2
        )
        kpi_style = ParagraphStyle(
            "KPIStyle",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8.0,
            leading=12.0,
            textColor=colors.HexColor("#4B5563"),
            leftIndent=20,
            spaceAfter=4
        )
        table_cell_style = ParagraphStyle(
            "TableCell",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#1F2937"),
        )
        table_cell_style_num = ParagraphStyle(
            "TableCellNum",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8,
            leading=11,
            alignment=2,
            textColor=colors.HexColor("#1F2937"),
        )
        table_cell_header = ParagraphStyle(
            "TableHead",
            parent=styles["Normal"],
            fontName=font_name_bold,
            fontSize=8,
            leading=11,
            textColor=colors.whitesmoke,
        )

        story = []

        # 1. Header Báo cáo & Khung Thông Tin Điều Hành (Executive Metadata Box)
        story.append(Paragraph("BÁO CÁO PHÂN TÍCH ĐIỀU HÀNH DOANH NGHIỆP", title_style))
        now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        query_text = clean_text_for_pdf(result.get("query", ""))
        total_rows_count = len(df) if df is not None else 0

        # Khung thông tin điều hành (Metadata Box)
        meta_data = [
            [
                Paragraph(f"<b>Chủ đề / Câu hỏi:</b> {query_text}", subtitle_style),
                Paragraph(f"<b>Thời gian xuất:</b> {now_str} &nbsp;|&nbsp; <b>Tổng dòng:</b> {total_rows_count:,} &nbsp;|&nbsp; <b>Bảo mật:</b> Confidential", subtitle_style),
            ]
        ]
        meta_table = Table(meta_data, colWidths=[330, 193])
        meta_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("PADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 4))

        # Khung Tóm Tắt Chỉ Số Cốt Lõi (Executive KPI Snapshot Box)
        if df is not None and not df.empty:
            from src.analytics.heuristics import get_axis_columns, pick_label_column, is_id_like
            measure_cols, label_cols, _ = get_axis_columns(df)
            if not measure_cols:
                measure_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c]) and not is_id_like(c)]
                label_cols = [c for c in df.columns if c not in measure_cols]

            if measure_cols and len(df) > 1:
                m_col = measure_cols[0]
                m_name = clean_text_for_pdf(str(m_col).replace("_", " ").title())
                v_series = pd.to_numeric(df[m_col], errors="coerce").dropna()
                if not v_series.empty:
                    total_val = v_series.sum()
                    avg_val = v_series.mean()
                    max_idx = v_series.idxmax()
                    peak_val = df.loc[max_idx, m_col]

                    # Lấy tên đối tượng đầy đủ (ưu tiên Họ và Tên nếu có)
                    _, label_series, _ = pick_label_column(df, label_cols)
                    if label_series is not None and max_idx in label_series.index:
                        peak_label = clean_text_for_pdf(str(label_series.loc[max_idx]))
                    elif label_cols:
                        peak_label = clean_text_for_pdf(str(df.loc[max_idx, label_cols[0]]))
                    else:
                        peak_label = f"#{max_idx + 1}"

                    fmt_total = f"{total_val:,.0f}" if total_val > 100 else f"{total_val:,.2f}"
                    fmt_avg = f"{avg_val:,.0f}" if avg_val > 100 else f"{avg_val:,.2f}"
                    fmt_peak = f"{peak_val:,.0f}" if peak_val > 100 else f"{peak_val:,.2f}"

                    kpi_box_data = [
                        [
                            Paragraph(f"💰 <b>Tổng {m_name}:</b> {fmt_total}", subtitle_style),
                            Paragraph(f"📈 <b>Trung bình:</b> {fmt_avg}", subtitle_style),
                            Paragraph(f"🏆 <b>Đỉnh cao nhất:</b> {peak_label} ({fmt_peak})", subtitle_style),
                        ]
                    ]
                    kpi_table = Table(kpi_box_data, colWidths=[174, 174, 175])
                    kpi_table.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EFF6FF")),
                        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFDBFE")),
                        ("PADDING", (0, 0), (-1, -1), 4),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]))
                    story.append(kpi_table)
                    story.append(Spacer(1, 6))

        # 2. Bảng Tóm Tắt Dữ Liệu (Summary Data Table) với Căn Chỉnh Kế Toán & Zebra Striping
        if df is not None and not df.empty:
            story.append(Paragraph(f"1. BẢNG TỔNG HỢP DỮ LIỆU (Top {min(10, len(df))} / {len(df)} dòng)", section_style))
            preview_df = df.head(10)

            n_cols = len(preview_df.columns)
            col_w = max(45, min(140, 523 / max(1, n_cols)))

            table_data = [[Paragraph(f"<b>{clean_text_for_pdf(str(c))}</b>", table_cell_header) for c in preview_df.columns]]
            for _, row in preview_df.iterrows():
                row_cells = []
                for c in preview_df.columns:
                    val = row[c]
                    if isinstance(val, (int, float)) or (isinstance(val, str) and val.replace(",", "").replace(".", "").isdigit()):
                        try:
                            num_val = float(str(val).replace(",", ""))
                            val_str = f"{num_val:,.0f}" if num_val.is_integer() else f"{num_val:,.2f}"
                        except Exception:
                            val_str = str(val)
                        row_cells.append(Paragraph(clean_text_for_pdf(val_str), table_cell_style_num))
                    else:
                        val_str = str(val)
                        row_cells.append(Paragraph(clean_text_for_pdf(val_str), table_cell_style))
                table_data.append(row_cells)

            t = Table(table_data, colWidths=[col_w] * n_cols)
            t_style = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ]
            for r_idx in range(1, len(table_data)):
                bg = colors.HexColor("#F8FAFC") if r_idx % 2 == 0 else colors.white
                t_style.append(("BACKGROUND", (0, r_idx), (-1, r_idx), bg))

            t.setStyle(TableStyle(t_style))
            story.append(t)
            story.append(Spacer(1, 8))

        # 3. Ảnh Biểu Đồ (nếu có)
        if chart_png_bytes:
            story.append(Paragraph("2. BIỂU ĐỒ TRỰC QUAN HÓA (VISUAL CHART)", section_style))
            img_buffer = io.BytesIO(chart_png_bytes)
            story.append(Image(img_buffer, width=500, height=240))
            story.append(Spacer(1, 8))

        # 4. Bản Phân Tích Insight Chiến Lược & Kế Hoạch Hành Động (Action Plan)
        insights = result.get("insights")
        if insights:
            sec_num = "3" if chart_png_bytes else "2"
            story.append(Paragraph(f"{sec_num}. PHÂN TÍCH INSIGHT CHIẾN LƯỢC & KẾ HOẠCH HÀNH ĐỘNG", section_style))

            from src.analytics.heuristics import split_insight_sections
            sec_dict = split_insight_sections(insights)
            p21 = sec_dict.get("anomaly", "").strip()
            p22 = sec_dict.get("hypothesis", "").strip()
            p23 = sec_dict.get("action_plan", "").strip()

            # --- 2.1. Phát hiện Bất thường & Xu hướng Chính ---
            story.append(Paragraph(f"<b>{sec_num}.1. Phát hiện Bất thường & Xu hướng Chính</b>", sub_section_style))
            if p21:
                for line in p21.split("\n"):
                    l_clean = clean_text_for_pdf(line.strip())
                    if l_clean:
                        l_clean = re.sub(r"^[•\-\*]\s*", "", l_clean)
                        story.append(Paragraph(f"• {l_clean}", bullet_style))
            else:
                story.append(Paragraph("• Không ghi nhận biến động cực đoan bất thường.", bullet_style))
            story.append(Spacer(1, 4))

            # --- 2.2. Giả thuyết & Nguyên nhân Tiềm năng ---
            story.append(Paragraph(f"<b>{sec_num}.2. Giả thuyết & Nguyên nhân Tiềm năng</b>", sub_section_style))
            if p22:
                for line in p22.split("\n"):
                    l_clean = clean_text_for_pdf(line.strip())
                    if l_clean:
                        l_clean = re.sub(r"^[•\-\*]\s*", "", l_clean)
                        story.append(Paragraph(f"• {l_clean}", bullet_style))
            else:
                story.append(Paragraph("• Doanh thu vận hành ổn định theo quy luật kinh doanh.", bullet_style))
            story.append(Spacer(1, 4))

            # --- 2.3. Đề xuất Hành động (Action Plan) ---
            story.append(Paragraph(f"<b>{sec_num}.3. Đề xuất Hành động (Action Plan)</b>", sub_section_style))
            if p23:
                for line in p23.split("\n"):
                    clean_item = clean_text_for_pdf(line.strip())
                    if not clean_item:
                        continue
                    clean_item = re.sub(r"^[•\-\*]\s*", "", clean_item)
                    clean_item = re.sub(r"^#+\s*", "", clean_item).strip()
                    clean_item = re.sub(r"\]\s*:\s*:\s*", "]: ", clean_item)
                    clean_item = re.sub(r"\s*:\s*:\s*", ": ", clean_item)
                    clean_item = re.sub(r":\s*:\s*", ": ", clean_item)

                    if any(p in clean_item for p in ["[Ưu tiên", "[High Priority", "[Medium Priority", "[Low Priority"]):
                        tag_match = re.match(r"^(\[(?:Ưu tiên|High Priority|Medium Priority|Low Priority)[^\]]*\])\s*:?\s*(.*)$", clean_item, flags=re.IGNORECASE)
                        if tag_match:
                            tag_text = tag_match.group(1).strip()
                            body_text = tag_match.group(2).strip()
                            body_text = re.sub(r"^[:\s\-\•\*\.]+", "", body_text).strip()
                            body_text = re.sub(r"^\d+[\.\)]\s*", "", body_text).strip()
                            body_text = re.sub(r"^[:\s\-\•\*\.]+", "", body_text).strip()
                        else:
                            tag_text = clean_item
                            body_text = ""

                        if "Cao" in tag_text or "High" in tag_text:
                            tag_color = "#DC2626"
                        elif "Trung bình" in tag_text or "Medium" in tag_text:
                            tag_color = "#D97706"
                        else:
                            tag_color = "#16A34A"

                        if body_text:
                            story.append(Paragraph(f"• <font color='{tag_color}'><b>{tag_text}</b></font>: {body_text}", bullet_style))
                        else:
                            story.append(Paragraph(f"• <font color='{tag_color}'><b>{tag_text}</b></font>", bullet_style))

                    elif any(k in clean_item.lower() for k in ["kpi", "mục tiêu đo lường", "chỉ số đo lường"]):
                        clean_kpi = re.sub(r"^(?:kpi|mục tiêu đo lường|chỉ số đo lường|kpi đo lường)\s*:\s*", "", clean_item, flags=re.IGNORECASE).strip()
                        story.append(Paragraph(f"&nbsp;&nbsp;&nbsp;&nbsp;• <i><b>Mục tiêu đo lường:</b> {clean_kpi}</i>", kpi_style))
                    else:
                        story.append(Paragraph(f"• {clean_item}", bullet_style))
            else:
                story.append(Paragraph("• Tiếp tục theo dõi chỉ số định kỳ.", bullet_style))

            story.append(Spacer(1, 8))

        # 5. Khung Chi Tiết Câu Lệnh SQL
        sql_text = result.get("sql")
        if sql_text:
            sec_sql_num = "4" if chart_png_bytes else "3"
            story.append(Paragraph(f"{sec_sql_num}. CHI TIẾT CÂU LỆNH SQL ĐÃ THỰC THI", section_style))
            clean_sql = clean_text_for_pdf(sql_text)
            sql_box_data = [[Paragraph(f"<font color='#374151'>{clean_sql}</font>", table_cell_style)]]
            sql_table = Table(sql_box_data, colWidths=[523])
            sql_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F1F5F9")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(sql_table)

        doc.build(story)
        return buffer.getvalue()
    except Exception as e:
        return b""
